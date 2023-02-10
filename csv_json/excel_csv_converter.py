'''
1) take path as cmd line arg
2) get all xlsx files in the dir
3) transform all spread sheets into csv files - one per sheet
4) name csv files <file_name>_<sheet_name>.csv
'''
from argparse import ArgumentParser
from sys import exit
from logging import debug, disable, basicConfig, DEBUG, CRITICAL
from pathlib import Path
from csv import writer
try:
    from openpyxl import load_workbook
except ImportError:
    exit('This program requires the openpyxl module to run.')
basicConfig(level=DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
disable(CRITICAL)

def parse_args():
    parser = ArgumentParser()
    parser.add_argument('dir', help='path to dir of xlsx files to convert')
    return parser.parse_args()

def get_files(location):
    files = []
    for file in Path(location).absolute().glob('*.xlsx'):
        debug(file)
        files.append(file)
    return files

def main():
    args = parse_args()

    files = get_files(args.dir)

    out_dir = Path(args.dir) / 'csvs'
    out_dir.mkdir(exist_ok=True, parents=True)

    for file in files:
        wb = load_workbook(str(file))
        sheet_names = wb.sheetnames
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]

            out_file_name = out_dir / (file.stem + '_' + sheet_name + '.csv')

            print('Creating {} from {} workbook {} sheet'.format(
                out_file_name, file.stem, sheet_name
            ))
            out_file = out_file_name.open('w')
            csv_writer = writer(out_file)

            for row_num in range(1, sheet.max_row + 1):
                row_data = []
                for col_num in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(row=row_num, column=col_num).value)
                csv_writer.writerow(row_data)
        
            out_file.close()

    print('Done')


if __name__ == '__main__':
    main()